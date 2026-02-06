#!/usr/bin/env python3
"""
Generate a complete, self-contained SQL migration file that:
1. Creates the entire database schema (enums, tables, functions, indexes, RLS, triggers)
2. Seeds all data from the Excel file (companies, equipment types, vehicles)

Usage:
  python scripts/generate_migration.py

Output:
  supabase/migrations/20260206_complete_migration.sql
"""
import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
SCHEMA_PATH = os.path.join(PROJECT_DIR, 'supabase', 'schema.sql')
SEED_TEMP = os.path.join(PROJECT_DIR, 'supabase', 'migrations', '_seed_temp.sql')
OUTPUT_PATH = os.path.join(PROJECT_DIR, 'supabase', 'migrations', '20260206_complete_migration.sql')

# ── Step 1: Read schema.sql ──────────────────────────────────────────────────
print("[1/4] Reading schema.sql ...")
with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
    schema_sql = f.read()

# ── Step 2: Generate seed SQL from Excel ─────────────────────────────────────
print("[2/4] Generating seed data from Excel ...")

import openpyxl
import uuid
import re
from datetime import datetime

EXCEL_PATH = r'C:\Users\Mark Ronnel Nieva\Desktop\Active Vehicless 2-1-2026 1-42-28 PM.xlsx'

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def sql_str(val):
    if val is None:
        return 'NULL'
    escaped = str(val).replace("'", "''")
    return f"'{escaped}'"

def classify_equipment(name):
    vehicles = ['Light Vehicle', 'Bus', 'Mini-Bus', 'Coach', 'Coaster', 'Ambulance',
                'Dyna', 'Mini Truck', 'Tanker', 'Sewage', 'Service Truck',
                'Flatbed Trailer', 'Concrete Mixer', 'Dump Truck', 'HIAB',
                'Concrete Pump']
    for v in vehicles:
        if v.lower() in str(name).lower():
            return 'vehicle'
    return 'heavy_equipment'

def extract_classification(name):
    match = re.search(r'\(([A-C])\)', str(name))
    return match.group(1) if match else None

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Active Vehicless']

companies = {}
equipment_types = {}

for r in range(2, ws.max_row + 1):
    company = clean(ws.cell(row=r, column=9).value)
    equipment = clean(ws.cell(row=r, column=10).value)
    if company and company not in companies:
        companies[company] = str(uuid.uuid4())
    if equipment and equipment not in equipment_types:
        equipment_types[equipment] = str(uuid.uuid4())

seed_lines = []
seed_lines.append("")
seed_lines.append("-- ============================================================================")
seed_lines.append("-- 11. SEED DATA - Companies")
seed_lines.append("-- ============================================================================")
seed_lines.append("")

for name, cid in companies.items():
    project = None
    if 'oxagon' in name.lower():
        project = 'OXAGON'
    elif 'sbc' in name.lower():
        project = 'SBC SITE'
    seed_lines.append(
        f"INSERT INTO companies (id, name, project, gate, is_active) "
        f"VALUES ({sql_str(cid)}, {sql_str(name)}, {sql_str(project)}, 'oxagon gate', true);"
    )

seed_lines.append("")
seed_lines.append("-- ============================================================================")
seed_lines.append("-- 12. SEED DATA - Equipment Types")
seed_lines.append("-- ============================================================================")
seed_lines.append("")

for name, eid in equipment_types.items():
    category = classify_equipment(name)
    classification = extract_classification(name)
    seed_lines.append(
        f"INSERT INTO equipment_types (id, name, category, classification, is_active) "
        f"VALUES ({sql_str(eid)}, {sql_str(name)}, {sql_str(category)}, {sql_str(classification)}, true);"
    )

seed_lines.append("")
seed_lines.append("-- ============================================================================")
seed_lines.append("-- 13. SEED DATA - Vehicles & Equipment (1,649 records)")
seed_lines.append("-- ============================================================================")
seed_lines.append("")

vehicle_count = 0
for r in range(2, ws.max_row + 1):
    plate = clean(ws.cell(row=r, column=4).value)
    if not plate:
        continue

    driver = clean(ws.cell(row=r, column=5).value)
    national_id = clean(ws.cell(row=r, column=6).value)
    company_name = clean(ws.cell(row=r, column=9).value)
    equipment_name = clean(ws.cell(row=r, column=10).value)
    project = clean(ws.cell(row=r, column=12).value)
    year = clean(ws.cell(row=r, column=13).value)
    next_insp = clean(ws.cell(row=r, column=15).value)
    status_raw = clean(ws.cell(row=r, column=16).value)
    blacklist = clean(ws.cell(row=r, column=17).value)
    gate = clean(ws.cell(row=r, column=11).value)

    company_id = companies.get(company_name) if company_name else None
    equip_id = equipment_types.get(equipment_name) if equipment_name else None

    # Normalize project name
    if project and project.lower() == 'oxagon':
        project = 'OXAGON'

    status_map = {
        'Verified': 'verified',
        'Inspection Overdue': 'inspection_overdue',
        'Updated Inspection Required': 'updated_inspection_required',
        'Rejected': 'rejected',
    }
    status = status_map.get(status_raw, 'verified')

    is_blacklisted = str(blacklist).lower() == 'yes' if blacklist else False
    if is_blacklisted:
        status = 'blacklisted'

    year_int = None
    if year:
        try:
            y = int(float(str(year)))
            # Fix data quality issue: year 20241 -> 2024
            if y > 2100:
                y = int(str(y)[:4])
            year_int = y
        except (ValueError, TypeError):
            pass

    next_date = 'NULL'
    if next_insp:
        try:
            if isinstance(next_insp, str):
                dt = datetime.fromisoformat(next_insp.replace(' ', 'T'))
            else:
                dt = next_insp
            next_date = sql_str(str(dt)[:10])
        except (ValueError, TypeError):
            pass

    vid = str(uuid.uuid4())
    year_sql = str(year_int) if year_int else 'NULL'

    seed_lines.append(
        f"INSERT INTO vehicles_equipment "
        f"(id, plate_number, driver_name, national_id, company_id, equipment_type_id, "
        f"year_of_manufacture, project, gate, status, next_inspection_date, blacklisted) "
        f"VALUES ({sql_str(vid)}, {sql_str(plate)}, {sql_str(driver)}, {sql_str(national_id)}, "
        f"{sql_str(company_id)}, {sql_str(equip_id)}, {year_sql}, {sql_str(project)}, "
        f"{sql_str(gate)}, {sql_str(status)}, {next_date}, {str(is_blacklisted).lower()});"
    )
    vehicle_count += 1

seed_sql = '\n'.join(seed_lines)

print(f"    -> {len(companies)} companies, {len(equipment_types)} equipment types, {vehicle_count} vehicles")

# ── Step 3: Combine into single migration ────────────────────────────────────
print("[3/4] Combining schema + seed into single migration file ...")

migration = f"""-- ============================================================================
-- VVS1 - COMPLETE DATABASE MIGRATION
-- Vehicle & Equipment Inspection Management System
-- ============================================================================
-- Generated: 2026-02-06
-- Source: Active Vehicless 2-1-2026 1-42-28 PM.xlsx
-- Records: {len(companies)} companies, {len(equipment_types)} equipment types, {vehicle_count} vehicles
-- ============================================================================
--
-- This is a SELF-CONTAINED migration file. Run it in the Supabase SQL Editor
-- to create the entire database schema and populate all data in one step.
--
-- Order of operations:
--   1.  Extensions (pgcrypto)
--   2.  Enum types (7)
--   3.  Tables (8) with foreign keys
--   4.  Helper functions (get_user_role, log_audit_event)
--   5.  Indexes (48)
--   6.  Row-Level Security enabled
--   7.  RLS Policies (42)
--   8.  Audit trigger function
--   9.  Audit triggers on tables
--   10. Updated_at auto-update triggers
--   11. Seed: Companies ({len(companies)})
--   12. Seed: Equipment Types ({len(equipment_types)})
--   13. Seed: Vehicles & Equipment ({vehicle_count})
-- ============================================================================

-- Wrap everything in a transaction for atomicity
BEGIN;

{schema_sql}

{seed_sql}

-- ============================================================================
-- 14. VERIFICATION QUERIES (run after migration to confirm)
-- ============================================================================

-- Uncomment these to verify the migration was successful:
-- SELECT 'companies' AS table_name, COUNT(*) AS row_count FROM companies
-- UNION ALL SELECT 'equipment_types', COUNT(*) FROM equipment_types
-- UNION ALL SELECT 'vehicles_equipment', COUNT(*) FROM vehicles_equipment;

COMMIT;

-- ============================================================================
-- END OF COMPLETE MIGRATION
-- ============================================================================
"""

# ── Step 4: Write output ─────────────────────────────────────────────────────
print(f"[4/4] Writing migration to: {OUTPUT_PATH}")
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write(migration)

file_size_kb = os.path.getsize(OUTPUT_PATH) / 1024
print(f"\nDone! Migration file created:")
print(f"  Path: {OUTPUT_PATH}")
print(f"  Size: {file_size_kb:.1f} KB")
print(f"  Data: {len(companies)} companies + {len(equipment_types)} equipment types + {vehicle_count} vehicles")
print(f"\nTo apply: Copy and paste into Supabase SQL Editor, or run via psql.")
