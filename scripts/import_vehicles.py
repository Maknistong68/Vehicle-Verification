#!/usr/bin/env python3
"""
Import vehicles from Excel file into Supabase-compatible SQL.
Run: python scripts/import_vehicles.py > supabase/seed.sql
"""
import openpyxl
import uuid
import re
from datetime import datetime

EXCEL_PATH = r'C:\Users\Mark Ronnel Nieva\Desktop\Active Vehicless 2-1-2026 1-42-28 PM.xlsx'

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def sql_str(val):
    if val is None:
        return 'NULL'
    escaped = str(val).replace("'", "''")
    return f"'{escaped}'"

def classify_equipment(name):
    vehicles = ['Light Vehicle', 'Bus', 'Mini-Bus', 'Coach', 'Coaster', 'Ambulance',
                'Dyna', 'Mini Truck', 'Tanker', 'Sewage', 'Service Truck',
                'Flatbed Trailer', 'Concrete Mixer', 'Dump Truck', 'HIAB']
    for v in vehicles:
        if v.lower() in str(name).lower():
            return 'vehicle'
    return 'heavy_equipment'

def extract_classification(name):
    match = re.search(r'\(([A-C])\)', str(name))
    return match.group(1) if match else None

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Active Vehicless']

    companies = {}
    equipment_types = {}

    for r in range(2, ws.max_row + 1):
        company = clean(ws.cell(row=r, column=9).value)
        equipment = clean(ws.cell(row=r, column=10).value)
        if company and company not in companies:
            companies[company] = str(uuid.uuid4())
        if equipment and equipment not in equipment_types:
            equipment_types[equipment] = str(uuid.uuid4())

    print("-- VVS1 Database Seed Script")
    print("-- Generated from Active Vehicles Excel file")
    print("-- Run this in Supabase SQL Editor after creating the schema")
    print()

    # Companies
    print("-- ============ COMPANIES ============")
    for name, cid in companies.items():
        project = None
        if 'oxagon' in name.lower():
            project = 'OXAGON'
        elif 'sbc' in name.lower():
            project = 'SBC SITE'
        print(f"INSERT INTO companies (id, name, project, gate, is_active) VALUES ({sql_str(cid)}, {sql_str(name)}, {sql_str(project)}, 'oxagon gate', true);")

    print()

    # Equipment Types
    print("-- ============ EQUIPMENT TYPES ============")
    for name, eid in equipment_types.items():
        category = classify_equipment(name)
        classification = extract_classification(name)
        print(f"INSERT INTO equipment_types (id, name, category, classification, is_active) VALUES ({sql_str(eid)}, {sql_str(name)}, {sql_str(category)}, {sql_str(classification)}, true);")

    print()

    # Vehicles/Equipment
    print("-- ============ VEHICLES & EQUIPMENT ============")
    count = 0
    for r in range(2, ws.max_row + 1):
        plate = clean(ws.cell(row=r, column=4).value)
        if not plate:
            continue

        driver = clean(ws.cell(row=r, column=5).value)
        national_id = clean(ws.cell(row=r, column=6).value)
        company_name = clean(ws.cell(row=r, column=9).value)
        equipment_name = clean(ws.cell(row=r, column=10).value)
        project = clean(ws.cell(row=r, column=12).value)
        year = clean(ws.cell(row=r, column=13).value)
        next_insp = clean(ws.cell(row=r, column=15).value)
        status_raw = clean(ws.cell(row=r, column=16).value)
        blacklist = clean(ws.cell(row=r, column=17).value)
        gate = clean(ws.cell(row=r, column=11).value)

        company_id = companies.get(company_name) if company_name else None
        equip_id = equipment_types.get(equipment_name) if equipment_name else None

        status_map = {
            'Verified': 'verified',
            'Inspection Overdue': 'inspection_overdue',
            'Updated Inspection Required': 'updated_inspection_required',
            'Rejected': 'rejected',
        }
        status = status_map.get(status_raw, 'verified')

        is_blacklisted = str(blacklist).lower() == 'yes' if blacklist else False
        if is_blacklisted:
            status = 'blacklisted'

        year_int = None
        if year:
            try:
                year_int = int(float(str(year)))
            except (ValueError, TypeError):
                pass

        next_date = 'NULL'
        if next_insp:
            try:
                if isinstance(next_insp, str):
                    dt = datetime.fromisoformat(next_insp.replace(' ', 'T'))
                else:
                    dt = next_insp
                next_date = sql_str(str(dt)[:10])
            except (ValueError, TypeError):
                pass

        vid = str(uuid.uuid4())
        year_sql = str(year_int) if year_int else 'NULL'

        print(f"INSERT INTO vehicles_equipment (id, plate_number, driver_name, national_id, company_id, equipment_type_id, year_of_manufacture, project, gate, status, next_inspection_date, blacklisted) VALUES ({sql_str(vid)}, {sql_str(plate)}, {sql_str(driver)}, {sql_str(national_id)}, {sql_str(company_id)}, {sql_str(equip_id)}, {year_sql}, {sql_str(project)}, {sql_str(gate)}, {sql_str(status)}, {next_date}, {str(is_blacklisted).lower()});")
        count += 1

    print()
    print(f"-- ============ DONE ============")
    print(f"-- Imported {len(companies)} companies, {len(equipment_types)} equipment types, and {count} vehicles/equipment")

if __name__ == '__main__':
    main()
